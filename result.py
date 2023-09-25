from openpyxl import load_workbook  
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
max_row=ws.max_row
count = 0
for row in range(2,max_row+1):
    hour=ws['B' + str(row)].value
    rate=ws['C' + str(row)].value
    if(type(hour)!=str and type(rate)!=str):
        salary=hour*rate
        if salary > 3000:
            count += 1
print(count)        
wb.close()
